# OP_CMS Reports API
# Story 4.4: Multi-dimensional Report Export

from sanic import Blueprint, json, request
from sanic.exceptions import NotFound, BadRequest
from typing import Optional
import logging
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import io

from backend.models.database_models import SettlementRecord, Customer, PriceConfig
from backend.dao.database_dao import DatabaseSessionFactory

logger = logging.getLogger(__name__)

reports_bp = Blueprint('reports', url_prefix='/reports')


@reports_bp.route('/export', methods=['POST'])
async def export_report(req: request.Request):
    """
    Export report to Excel/CSV/PDF
    
    Request Body:
    {
        "report_type": "customer_analysis",  // customer_analysis, revenue, collection
        "format": "excel",  // excel, csv, pdf
        "filters": {
            "date_from": "2026-01-01",
            "date_to": "2026-02-28",
            "customer_ids": [1, 2, 3]
        }
    }
    
    Returns:
    {
        "success": true,
        "data": {
            "file_content": "base64_encoded_content",
            "file_name": "report_20260228.xlsx",
            "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    }
    """
    try:
        data = req.json
        
        # Validate required fields
        report_type = data.get('report_type')
        export_format = data.get('format', 'excel')
        filters = data.get('filters', {})
        
        if not report_type:
            return json({
                'success': False,
                'error': 'Missing required fields',
                'message': 'report_type is required'
            }, status=400)
        
        # Get database session
        session_factory = DatabaseSessionFactory()
        session = session_factory.get_session()
        
        try:
            # Generate report data based on type
            if report_type == 'customer_analysis':
                report_data = generate_customer_analysis_report(session, filters)
            elif report_type == 'revenue':
                report_data = generate_revenue_report(session, filters)
            elif report_type == 'collection':
                report_data = generate_collection_report(session, filters)
            else:
                return json({
                    'success': False,
                    'error': 'Invalid report type',
                    'message': f'Unsupported report type: {report_type}'
                }, status=400)
            
            # Export to requested format
            if export_format == 'excel':
                file_content, file_name = export_to_excel(report_data, report_type)
                file_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif export_format == 'csv':
                file_content, file_name = export_to_csv(report_data, report_type)
                file_type = 'text/csv'
            elif export_format == 'pdf':
                # For PDF, we'll return a simple text representation
                # In production, use reportlab or WeasyPrint
                file_content, file_name = export_to_pdf_text(report_data, report_type)
                file_type = 'application/pdf'
            else:
                return json({
                    'success': False,
                    'error': 'Invalid format',
                    'message': f'Unsupported export format: {export_format}'
                }, status=400)
            
            # Encode to base64 for transmission
            import base64
            encoded_content = base64.b64encode(file_content).decode('utf-8')
            
            return json({
                'success': True,
                'data': {
                    'file_content': encoded_content,
                    'file_name': file_name,
                    'file_type': file_type,
                    'file_size': len(file_content),
                    'report_type': report_type,
                    'export_format': export_format
                },
                'message': f'Report exported successfully in {export_format} format'
            })
            
        finally:
            session.close()
            
    except Exception as e:
        logger.error(f"Failed to export report: {str(e)}")
        return json({
            'success': False,
            'error': 'Internal server error',
            'message': str(e)
        }, status=500)


def generate_customer_analysis_report(session, filters: dict) -> dict:
    """Generate customer analysis report data"""
    # Get date range
    date_from = datetime.fromisoformat(filters.get('date_from')) if filters.get('date_from') else datetime.utcnow() - timedelta(days=30)
    date_to = datetime.fromisoformat(filters.get('date_to')) if filters.get('date_to') else datetime.utcnow()
    
    # Get customers
    customer_ids = filters.get('customer_ids')
    query = session.query(Customer)
    if customer_ids:
        query = query.filter(Customer.id.in_(customer_ids))
    customers = query.all()
    
    # Generate report data
    report_data = {
        'title': '客户分析报表',
        'generated_at': datetime.utcnow().isoformat(),
        'date_range': {
            'from': date_from.isoformat(),
            'to': date_to.isoformat()
        },
        'summary': {
            'total_customers': len(customers),
            'active_customers': sum(1 for c in customers if c.status == 'active'),
            'vip_customers': sum(1 for c in customers if c.level == 'vip')
        },
        'customers': []
    }
    
    for customer in customers:
        # Get customer settlements
        settlements = session.query(SettlementRecord).filter(
            SettlementRecord.customer_id == customer.id,
            SettlementRecord.created_at >= date_from,
            SettlementRecord.created_at <= date_to
        ).all()
        
        total_revenue = sum([s.total_amount for s in settlements]) if settlements else Decimal('0')
        total_usage = sum([s.usage_quantity for s in settlements]) if settlements else Decimal('0')
        
        report_data['customers'].append({
            'customer_id': customer.customer_id,
            'company_name': customer.company_name,
            'contact_name': customer.contact_name,
            'contact_phone': customer.contact_phone,
            'level': customer.level,
            'status': customer.status,
            'total_revenue': float(total_revenue),
            'total_usage': float(total_usage) if total_usage else 0,
            'settlement_count': len(settlements)
        })
    
    return report_data


def generate_revenue_report(session, filters: dict) -> dict:
    """Generate revenue report data"""
    date_from = datetime.fromisoformat(filters.get('date_from')) if filters.get('date_from') else datetime.utcnow() - timedelta(days=30)
    date_to = datetime.fromisoformat(filters.get('date_to')) if filters.get('date_to') else datetime.utcnow()
    
    # Get settlements
    settlements = session.query(SettlementRecord).filter(
        SettlementRecord.created_at >= date_from,
        SettlementRecord.created_at <= date_to
    ).all()
    
    # Calculate revenue by status
    revenue_by_status = {}
    for settlement in settlements:
        status = settlement.status
        if status not in revenue_by_status:
            revenue_by_status[status] = Decimal('0')
        revenue_by_status[status] += settlement.total_amount
    
    report_data = {
        'title': '收入报表',
        'generated_at': datetime.utcnow().isoformat(),
        'date_range': {
            'from': date_from.isoformat(),
            'to': date_to.isoformat()
        },
        'summary': {
            'total_revenue': float(sum(revenue_by_status.values())),
            'paid_revenue': float(revenue_by_status.get('paid', Decimal('0'))),
            'pending_revenue': float(revenue_by_status.get('pending', Decimal('0'))),
            'approved_revenue': float(revenue_by_status.get('approved', Decimal('0')))
        },
        'settlements': []
    }
    
    for settlement in settlements:
        report_data['settlements'].append({
            'record_id': settlement.record_id,
            'customer_id': settlement.customer_id,
            'period_start': settlement.period_start.isoformat() if settlement.period_start else None,
            'period_end': settlement.period_end.isoformat() if settlement.period_end else None,
            'usage_quantity': float(settlement.usage_quantity) if settlement.usage_quantity else 0,
            'unit_price': float(settlement.unit_price) if settlement.unit_price else 0,
            'total_amount': float(settlement.total_amount) if settlement.total_amount else 0,
            'status': settlement.status,
            'created_at': settlement.created_at.isoformat() if settlement.created_at else None
        })
    
    return report_data


def generate_collection_report(session, filters: dict) -> dict:
    """Generate collection report data"""
    date_from = datetime.fromisoformat(filters.get('date_from')) if filters.get('date_from') else datetime.utcnow() - timedelta(days=30)
    date_to = datetime.fromisoformat(filters.get('date_to')) if filters.get('date_to') else datetime.utcnow()
    
    # Get settlements
    settlements = session.query(SettlementRecord).filter(
        SettlementRecord.created_at >= date_from,
        SettlementRecord.created_at <= date_to
    ).all()
    
    # Calculate collection metrics
    total_amount = sum([s.total_amount for s in settlements])
    paid_amount = sum([s.total_amount for s in settlements if s.status == 'paid'])
    pending_amount = sum([s.total_amount for s in settlements if s.status in ['pending', 'approved']])
    
    collection_rate = paid_amount / total_amount if total_amount > 0 else Decimal('0')
    
    report_data = {
        'title': '回款报表',
        'generated_at': datetime.utcnow().isoformat(),
        'date_range': {
            'from': date_from.isoformat(),
            'to': date_to.isoformat()
        },
        'summary': {
            'total_amount': float(total_amount),
            'paid_amount': float(paid_amount),
            'pending_amount': float(pending_amount),
            'collection_rate': float(collection_rate)
        },
        'settlements': []
    }
    
    for settlement in settlements:
        report_data['settlements'].append({
            'record_id': settlement.record_id,
            'customer_id': settlement.customer_id,
            'period_start': settlement.period_start.isoformat() if settlement.period_start else None,
            'period_end': settlement.period_end.isoformat() if settlement.period_end else None,
            'total_amount': float(settlement.total_amount) if settlement.total_amount else 0,
            'status': settlement.status,
            'created_at': settlement.created_at.isoformat() if settlement.created_at else None
        })
    
    return report_data


def export_to_excel(report_data: dict, report_type: str):
    """Export report data to Excel format"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_type
        
        # Add title
        ws.cell(row=1, column=1, value=report_data['title'])
        ws.cell(row=2, column=1, value=f"生成时间：{report_data['generated_at']}")
        
        # Add data based on report type
        if report_type == 'customer_analysis':
            # Add headers
            headers = ['客户 ID', '公司名称', '联系人', '联系电话', '等级', '状态', '总收入', '总用量', '结算次数']
            for col, header in enumerate(headers, 1):
                ws.cell(row=4, column=col, value=header)
            
            # Add data
            for row_idx, customer in enumerate(report_data['customers'], 5):
                ws.cell(row=row_idx, column=1, value=customer['customer_id'])
                ws.cell(row=row_idx, column=2, value=customer['company_name'])
                ws.cell(row=row_idx, column=3, value=customer['contact_name'])
                ws.cell(row=row_idx, column=4, value=customer['contact_phone'])
                ws.cell(row=row_idx, column=5, value=customer['level'])
                ws.cell(row=row_idx, column=6, value=customer['status'])
                ws.cell(row=row_idx, column=7, value=customer['total_revenue'])
                ws.cell(row=row_idx, column=8, value=customer['total_usage'])
                ws.cell(row=row_idx, column=9, value=customer['settlement_count'])
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_name = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return output.getvalue(), file_name
        
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_to_csv(report_data, report_type)


def export_to_csv(report_data: dict, report_type: str):
    """Export report data to CSV format"""
    output = io.StringIO()
    
    if report_type == 'customer_analysis' and report_data.get('customers'):
        # Write CSV
        writer = csv.writer(output)
        
        # Add headers
        writer.writerow(['客户 ID', '公司名称', '联系人', '联系电话', '等级', '状态', '总收入', '总用量', '结算次数'])
        
        # Add data
        for customer in report_data['customers']:
            writer.writerow([
                customer['customer_id'],
                customer['company_name'],
                customer['contact_name'],
                customer['contact_phone'],
                customer['level'],
                customer['status'],
                customer['total_revenue'],
                customer['total_usage'],
                customer['settlement_count']
            ])
    
    file_name = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return output.getvalue().encode('utf-8'), file_name


def export_to_pdf_text(report_data: dict, report_type: str):
    """Export report to PDF (simple text format - placeholder for production PDF)"""
    # In production, use reportlab or WeasyPrint for proper PDF generation
    output = io.StringIO()
    
    output.write(f"{report_data['title']}\n")
    output.write(f"生成时间：{report_data['generated_at']}\n")
    output.write("=" * 50 + "\n\n")
    
    if 'summary' in report_data:
        output.write("汇总:\n")
        for key, value in report_data['summary'].items():
            output.write(f"  {key}: {value}\n")
        output.write("\n")
    
    file_name = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return output.getvalue().encode('utf-8'), file_name
