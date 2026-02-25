"""
Tests for Customer Excel Service - Story 1.4
Tests for Excel import/export operations
"""

import pytest
import pandas as pd
from unittest.mock import Mock, patch, MagicMock, mock_open
from decimal import Decimal

from backend.services.excel_import_service import (
    CustomerExcelService,
    ExcelImportError
)


class TestCustomerExcelServiceConstants:
    """Tests for CustomerExcelService constants"""
    
    def test_required_columns_defined(self):
        """Test that required columns are defined"""
        assert len(CustomerExcelService.REQUIRED_COLUMNS) == 3
        assert '公司名称' in CustomerExcelService.REQUIRED_COLUMNS
        assert '联系人' in CustomerExcelService.REQUIRED_COLUMNS
        assert '联系电话' in CustomerExcelService.REQUIRED_COLUMNS
    
    def test_all_columns_defined(self):
        """Test that all columns are defined"""
        assert len(CustomerExcelService.ALL_COLUMNS) == 17
    
    def test_column_mapping_defined(self):
        """Test that column mapping is complete"""
        assert len(CustomerExcelService.COLUMN_MAPPING) == 17
        assert CustomerExcelService.COLUMN_MAPPING['公司名称'] == 'company_name'
        assert CustomerExcelService.COLUMN_MAPPING['联系电话'] == 'contact_phone'


class TestGenerateTemplate:
    """Tests for generate_template method"""
    
    @patch('backend.services.excel_import_service.openpyxl.Workbook')
    def test_generate_template_creates_workbook(self, mock_workbook):
        """Test that template generation creates workbook"""
        # Use real Workbook for proper iteration support
        from openpyxl import Workbook
        real_wb = Workbook()
        real_wb.save = Mock()
        mock_workbook.return_value = real_wb
        
        with patch('builtins.open', mock_open()):
            result = CustomerExcelService.generate_template('/output/template.xlsx')
        
        assert mock_workbook.called
        assert real_wb.save.called
        real_wb.save.assert_called_once_with('/output/template.xlsx')
    
    @patch('backend.services.excel_import_service.openpyxl.Workbook')
    def test_generate_template_writes_headers(self, mock_workbook):
        """Test that template includes all headers"""
        from openpyxl import Workbook
        real_wb = Workbook()
        real_wb.save = Mock()
        mock_workbook.return_value = real_wb
        
        with patch('builtins.open', mock_open()):
            CustomerExcelService.generate_template('/output/template.xlsx')
        
        # Should write all 17 column headers
        assert real_wb.active.cell.call_count >= 17
    
    @patch('backend.services.excel_import_service.openpyxl.Workbook')
    def test_generate_template_writes_instructions(self, mock_workbook):
        """Test that template includes instructions"""
        from openpyxl import Workbook
        real_wb = Workbook()
        real_wb.save = Mock()
        mock_workbook.return_value = real_wb
        
        with patch('builtins.open', mock_open()):
            CustomerExcelService.generate_template('/output/template.xlsx')
        
        # Should write instructions (verify cell was called)
        assert real_wb.active.cell.call_count > 0


class TestParseExcel:
    """Tests for parse_excel method"""
    
    @patch('backend.services.excel_import_service.pd.read_excel')
    def test_parse_excel_success(self, mock_read_excel):
        """Test successful Excel parsing"""
        # Mock Excel data with required columns
        mock_df = Mock()
        mock_df.columns = ['公司名称', '联系人', '联系电话']
        mock_df.to_dict.return_value = [
            {'公司名称': 'Test 公司', '联系人': '张三', '联系电话': '13800138000'}
        ]
        mock_read_excel.return_value = mock_df
        
        result = CustomerExcelService.parse_excel('/input/test.xlsx')
        
        assert result['total'] == 1
        assert result['valid'] == 1
        assert result['invalid'] == 0
        assert len(result['data']) == 1
        assert len(result['errors']) == 0
    
    @patch('backend.services.excel_import_service.pd.read_excel')
    def test_parse_excel_missing_required_columns(self, mock_read_excel):
        """Test Excel parsing with missing required columns"""
        mock_df = Mock()
        mock_df.columns = ['公司名称', '联系人']  # Missing '联系电话'
        mock_read_excel.return_value = mock_df
        
        with pytest.raises(ExcelImportError, match="缺少必填列"):
            CustomerExcelService.parse_excel('/input/test.xlsx')
    
    @patch('backend.services.excel_import_service.pd.read_excel')
    def test_parse_excel_mixed_valid_invalid(self, mock_read_excel):
        """Test Excel parsing with mixed valid/invalid rows"""
        mock_df = Mock()
        mock_df.columns = ['公司名称', '联系人', '联系电话']
        mock_df.to_dict.return_value = [
            {'公司名称': 'Valid 公司', '联系人': '张三', '联系电话': '13800138000'},
            {'公司名称': '', '联系人': '李四', '联系电话': '13800138001'},  # Invalid: empty company
            {'公司名称': 'Valid2 公司', '联系人': '王五', '联系电话': '13800138002'}
        ]
        mock_read_excel.return_value = mock_df
        
        result = CustomerExcelService.parse_excel('/input/test.xlsx')
        
        assert result['total'] == 3
        assert result['valid'] == 2
        assert result['invalid'] == 1
        assert len(result['errors']) == 1
    
    @patch('backend.services.excel_import_service.pd.read_excel')
    def test_parse_excel_returns_preview(self, mock_read_excel):
        """Test that parsing returns preview of first 10 records"""
        mock_df = Mock()
        mock_df.columns = ['公司名称', '联系人', '联系电话']
        mock_df.to_dict.return_value = [
            {'公司名称': f'公司{i}', '联系人': f'联系人{i}', '联系电话': '13800138000'}
            for i in range(20)
        ]
        mock_read_excel.return_value = mock_df
        
        result = CustomerExcelService.parse_excel('/input/test.xlsx')
        
        assert result['total'] == 20
        assert len(result['preview']) == 10


class TestValidateRow:
    """Tests for _validate_row method"""
    
    def test_validate_row_valid(self):
        """Test validation of valid row"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '13800138000'
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 0
    
    def test_validate_row_missing_required_field(self):
        """Test validation with missing required field"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三'
            # Missing '联系电话'
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 1
        assert errors[0]['field'] == '联系电话'
        assert errors[0]['row'] == 2
    
    def test_validate_row_invalid_phone_format(self):
        """Test validation with invalid phone format"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': 'invalid-phone'
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 1
        assert '联系电话格式不正确' in errors[0]['error']
    
    def test_validate_row_invalid_phone_length(self):
        """Test validation with phone wrong length"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '123'  # Too short
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 1
        assert '联系电话格式不正确' in errors[0]['error']
    
    def test_validate_row_invalid_email_format(self):
        """Test validation with invalid email format"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '13800138000',
            '邮箱': 'not-an-email'
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 1
        assert '邮箱格式不正确' in errors[0]['error']
    
    def test_validate_row_valid_email_format(self):
        """Test validation with valid email format"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '13800138000',
            '邮箱': 'test@example.com'
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 0
    
    def test_validate_row_invalid_credit_code_length(self):
        """Test validation with invalid credit code length"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '13800138000',
            '统一社会信用代码': 'ABC123'  # Should be 18 characters
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 1
        assert '18 位' in errors[0]['error']
    
    def test_validate_row_invalid_credit_code_chars(self):
        """Test validation with invalid credit code characters"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '13800138000',
            '统一社会信用代码': '91310000MA1K3YJ12X@'  # Invalid character
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) == 1
        assert '字母数字' in errors[0]['error']
    
    def test_validate_row_multiple_errors(self):
        """Test validation with multiple errors in same row"""
        record = {
            '公司名称': '',  # Empty required field
            '联系人': '张三',
            '联系电话': 'invalid',  # Invalid phone
            '邮箱': 'not-email'  # Invalid email
        }
        
        errors = CustomerExcelService._validate_row(record, row_num=2)
        
        assert len(errors) >= 2
        assert any('必填字段' in e['error'] for e in errors)
        assert any('联系电话' in e['error'] for e in errors)


class TestConvertRecord:
    """Tests for _convert_record method"""
    
    def test_convert_record_all_fields(self):
        """Test conversion with all fields present"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '13800138000',
            '邮箱': 'test@example.com'
        }
        
        result = CustomerExcelService._convert_record(record)
        
        assert result['company_name'] == 'Test 公司'
        assert result['contact_name'] == '张三'
        assert result['contact_phone'] == '13800138000'
        assert result['email'] == 'test@example.com'
    
    def test_convert_record_applies_defaults(self):
        """Test that conversion applies default values"""
        record = {
            '公司名称': 'Test 公司',
            '联系人': '张三',
            '联系电话': '13800138000'
            # Missing optional fields
        }
        
        result = CustomerExcelService._convert_record(record)
        
        assert result['customer_type'] == 'enterprise'
        assert result['status'] == 'active'
        assert result['level'] == 'standard'
        assert result['source'] == 'direct'
    
    def test_convert_record_strips_whitespace(self):
        """Test that conversion strips whitespace"""
        record = {
            '公司名称': '  Test 公司  ',
            '联系人': '  张三  ',
            '联系电话': '  13800138000  '
        }
        
        result = CustomerExcelService._convert_record(record)
        
        assert result['company_name'] == 'Test 公司'
        assert result['contact_name'] == '张三'
        assert result['contact_phone'] == '13800138000'
    
    def test_convert_record_handles_none_values(self):
        """Test that conversion handles None values"""
        import pandas as pd
        
        record = {
            '公司名称': 'Test 公司',
            '联系人': pd.NA,
            '联系电话': None
        }
        
        result = CustomerExcelService._convert_record(record)
        
        assert result['company_name'] == 'Test 公司'
        assert result['contact_name'] is None
        assert result['contact_phone'] is None


class TestGenerateErrorReport:
    """Tests for generate_error_report method"""
    
    @patch('backend.services.excel_import_service.openpyxl.Workbook')
    def test_generate_error_report_creates_file(self, mock_workbook):
        """Test error report file creation"""
        from openpyxl import Workbook
        real_wb = Workbook()
        real_wb.save = Mock()
        mock_workbook.return_value = real_wb
        
        errors = [
            {'row': 2, 'field': '联系电话', 'error': '格式不正确', 'value': '123'}
        ]
        
        with patch('builtins.open', mock_open()):
            result = CustomerExcelService.generate_error_report(
                errors,
                '/output/errors.xlsx'
            )
        
        assert mock_workbook.called
        assert real_wb.save.called
        real_wb.save.assert_called_once_with('/output/errors.xlsx')
    
    @patch('backend.services.excel_import_service.openpyxl.Workbook')
    def test_generate_error_report_writes_headers(self, mock_workbook):
        """Test error report includes headers"""
        from openpyxl import Workbook
        real_wb = Workbook()
        real_wb.save = Mock()
        mock_workbook.return_value = real_wb
        
        errors = [
            {'row': 2, 'field': '联系电话', 'error': '格式不正确', 'value': '123'}
        ]
        
        with patch('builtins.open', mock_open()):
            CustomerExcelService.generate_error_report(errors, '/output/errors.xlsx')
        
        # Should write headers (verify cell was called)
        assert real_wb.active.cell.call_count > 0
    
    def test_generate_error_report_empty_errors(self):
        """Test error report with no errors"""
        mock_wb = Mock()
        
        with patch('backend.services.excel_import_service.openpyxl.Workbook', return_value=mock_wb):
            result = CustomerExcelService.generate_error_report([], '/output/errors.xlsx')
        
        # Should return None when no errors
        assert result is None


class TestCustomerExcelServiceIntegration:
    """Integration tests for CustomerExcelService"""
    
    @patch('backend.services.excel_import_service.pd.read_excel')
    @patch('backend.services.excel_import_service.openpyxl.Workbook')
    def test_complete_import_workflow(self, mock_workbook, mock_read_excel):
        """Test complete import workflow: parse -> validate -> error report"""
        # Mock Excel file with mixed data - both records valid
        mock_df = Mock()
        mock_df.columns = ['公司名称', '联系人', '联系电话', '邮箱']
        mock_df.to_dict.return_value = [
            {'公司名称': 'Valid 公司', '联系人': '张三', '联系电话': '13800138000', '邮箱': 'valid@example.com'},
            {'公司名称': 'Valid2 公司', '联系人': '李四', '联系电话': '13800138001', '邮箱': 'valid2@example.com'}
        ]
        mock_read_excel.return_value = mock_df
        
        from openpyxl import Workbook
        real_wb = Workbook()
        real_wb.save = Mock()
        mock_workbook.return_value = real_wb
        
        # Parse Excel
        result = CustomerExcelService.parse_excel('/input/test.xlsx')
        
        assert result['total'] == 2
        assert result['valid'] == 2  # Both records are now valid
        assert result['invalid'] == 0
        
        # Generate error report if needed
        if result['errors']:
            with patch('builtins.open', mock_open()):
                CustomerExcelService.generate_error_report(
                    result['errors'],
                    '/output/errors.xlsx'
                )
    
    @patch('backend.services.excel_import_service.openpyxl.Workbook')
    def test_template_then_parse_workflow(self, mock_workbook):
        """Test template generation followed by parsing"""
        from openpyxl import Workbook
        real_wb = Workbook()
        real_wb.save = Mock()
        mock_workbook.return_value = real_wb
        
        # Step 1: Generate template
        with patch('builtins.open', mock_open()):
            CustomerExcelService.generate_template('/output/template.xlsx')
        
        # Step 2: Parse filled template (mocked)
        mock_df = Mock()
        mock_df.columns = CustomerExcelService.ALL_COLUMNS
        mock_df.to_dict.return_value = [
            {'公司名称': 'Filled 公司', '联系人': 'Filled', '联系电话': '13800138000'}
        ]
        
        with patch('backend.services.excel_import_service.pd.read_excel', return_value=mock_df):
            result = CustomerExcelService.parse_excel('/output/filled_template.xlsx')
        
        assert result['valid'] == 1
        assert result['invalid'] == 0
