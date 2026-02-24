# OP_CMS Excel Import Service
# Story 1.4: Customer Data Batch Import

import openpyxl
import pandas as pd
from typing import List, Dict, Any, Optional
from datetime import datetime
import uuid
import re
from decimal import Decimal

from backend.models.database_models import CustomerCreate


class ExcelImportError(Exception):
    """Custom exception for Excel import errors"""
    def __init__(self, message: str, row: Optional[int] = None):
        super().__init__(message)
        self.row = row


class CustomerExcelService:
    """Service for handling Excel import/export operations"""
    
    # Required columns in Excel template
    REQUIRED_COLUMNS = [
        '公司名称',
        '联系人',
        '联系电话'
    ]
    
    # All supported columns
    ALL_COLUMNS = [
        '公司名称', '联系人', '联系电话', '统一社会信用代码', '客户类型',
        '省份', '城市', '地址', '邮箱', '网站', '所属行业',
        'ERP 系统', 'ERP 客户代码', '客户状态', '客户等级', '来源渠道', '备注'
    ]
    
    # Column mapping to model fields
    COLUMN_MAPPING = {
        '公司名称': 'company_name',
        '联系人': 'contact_name',
        '联系电话': 'contact_phone',
        '统一社会信用代码': 'credit_code',
        '客户类型': 'customer_type',
        '省份': 'province',
        '城市': 'city',
        '地址': 'address',
        '邮箱': 'email',
        '网站': 'website',
        '所属行业': 'industry',
        'ERP 系统': 'erp_system',
        'ERP 客户代码': 'erp_customer_code',
        '客户状态': 'status',
        '客户等级': 'level',
        '来源渠道': 'source',
        '备注': 'remarks'
    }
    
    @classmethod
    def generate_template(cls, output_path: str):
        """
        Generate Excel template file
        
        Args:
            output_path: Path to save template file
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "客户导入模板"
        
        # Add headers
        headers = cls.ALL_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add example data
        example_data = [
            ['Test 公司', '张三', '13800138000', '91310000MA1K3YJ12X', 'enterprise',
             'Shanghai', 'Shanghai', 'Test Road 123', 'test@example.com', 'https://test.com',
             'Technology', 'SAP', 'C001', 'active', 'vip', 'direct', 'Example customer'],
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add instructions
        ws.cell(row=1, column=len(headers) + 1, value="填写说明：")
        ws.cell(row=2, column=len(headers) + 1, value="1. 必填字段：公司名称、联系人、联系电话")
        ws.cell(row=3, column=len(headers) + 1, value="2. 联系电话格式：11 位数字或 +86 开头")
        ws.cell(row=4, column=len(headers) + 1, value="3. 统一社会信用代码：18 位字母数字")
        ws.cell(row=5, column=len(headers) + 1, value="4. 邮箱格式：name@domain.com")
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    @classmethod
    def parse_excel(cls, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel file and return data with validation
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with parsed data, errors, and statistics
        """
        try:
            # Read Excel file with pandas
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Check for required columns
            missing_columns = []
            for col in cls.REQUIRED_COLUMNS:
                if col not in df.columns:
                    missing_columns.append(col)
            
            if missing_columns:
                raise ExcelImportError(f"缺少必填列：{', '.join(missing_columns)}")
            
            # Convert to list of dicts
            records = df.to_dict('records')
            
            # Validate each row
            errors = []
            valid_records = []
            
            for idx, record in enumerate(records, 2):  # Start from 2 (header is row 1)
                row_errors = cls._validate_row(record, idx)
                
                if row_errors:
                    errors.extend(row_errors)
                else:
                    valid_records.append(cls._convert_record(record))
            
            return {
                'total': len(records),
                'valid': len(valid_records),
                'invalid': len(errors),
                'data': valid_records,
                'errors': errors,
                'preview': valid_records[:10]  # First 10 valid records
            }
            
        except pd.ExcelError as e:
            raise ExcelImportError(f"Excel 文件解析失败：{str(e)}")
        except Exception as e:
            raise ExcelImportError(f"文件处理失败：{str(e)}")
    
    @classmethod
    def _validate_row(cls, record: Dict[str, Any], row_num: int) -> List[Dict[str, Any]]:
        """
        Validate a single row of data
        
        Args:
            record: Row data as dictionary
            row_num: Row number for error reporting
            
        Returns:
            List of validation errors
        """
        errors = []
        
        # Check required fields
        for field in cls.REQUIRED_COLUMNS:
            value = record.get(field)
            if pd.isna(value) or (isinstance(value, str) and not value.strip()):
                errors.append({
                    'row': row_num,
                    'field': field,
                    'error': f'{field} 为必填字段',
                    'value': None
                })
        
        # Validate phone format
        phone = record.get('联系电话')
        if phone and not pd.isna(phone):
            phone_str = str(phone).strip()
            cleaned = re.sub(r'[\+\-\(\)\s]', '', phone_str)
            if not cleaned.isdigit() or len(cleaned) < 8 or len(cleaned) > 15:
                errors.append({
                    'row': row_num,
                    'field': '联系电话',
                    'error': '联系电话格式不正确（应为 8-15 位数字）',
                    'value': phone_str
                })
        
        # Validate email format
        email = record.get('邮箱')
        if email and not pd.isna(email):
            email_str = str(email).strip()
            pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(pattern, email_str):
                errors.append({
                    'row': row_num,
                    'field': '邮箱',
                    'error': '邮箱格式不正确',
                    'value': email_str
                })
        
        # Validate credit code
        credit_code = record.get('统一社会信用代码')
        if credit_code and not pd.isna(credit_code):
            code_str = str(credit_code).strip()
            if len(code_str) != 18 or not code_str.isalnum():
                errors.append({
                    'row': row_num,
                    'field': '统一社会信用代码',
                    'error': '统一社会信用代码格式不正确（应为 18 位字母数字）',
                    'value': code_str
                })
        
        return errors
    
    @classmethod
    def _convert_record(cls, record: Dict[str, Any]) -> Dict[str, Any]:
        """
        Convert Excel record to CustomerCreate format
        
        Args:
            record: Raw Excel record
            
        Returns:
            Dictionary ready for CustomerCreate
        """
        converted = {}
        
        for excel_col, model_field in cls.COLUMN_MAPPING.items():
            value = record.get(excel_col)
            
            if pd.isna(value):
                value = None
            elif isinstance(value, str):
                value = value.strip() if value else None
            
            converted[model_field] = value
        
        # Set defaults for None values
        if converted.get('customer_type') is None:
            converted['customer_type'] = 'enterprise'
        if converted.get('status') is None:
            converted['status'] = 'active'
        if converted.get('level') is None:
            converted['level'] = 'standard'
        if converted.get('source') is None:
            converted['source'] = 'direct'
        
        return converted
    
    @classmethod
    def generate_error_report(cls, errors: List[Dict[str, Any]], output_path: str):
        """
        Generate Excel report with errors
        
        Args:
            errors: List of error dictionaries
            output_path: Path to save error report
        """
        if not errors:
            return
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "导入错误报告"
        
        # Headers
        headers = ['行号', '字段', '错误信息', '原始值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        # Error rows
        for idx, error in enumerate(errors, 2):
            ws.cell(row=idx, column=1, value=error['row'])
            ws.cell(row=idx, column=2, value=error['field'])
            ws.cell(row=idx, column=3, value=error['error'])
            ws.cell(row=idx, column=4, value=error.get('value', ''))
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save
        wb.save(output_path)
        return output_path
